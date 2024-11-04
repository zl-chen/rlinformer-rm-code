# %%
from models.ddpg import Actor,Critic
import numpy as np
import torch 
import os
import time
from collections import Counter
from scipy.special import softmax
from sktime.performance_metrics.forecasting import mean_absolute_error,mean_absolute_percentage_error,mean_squared_error
from sklearn.metrics import r2_score
import torch
import torch.nn as nn 
import torch.nn.functional as F
import matplotlib.pyplot as plt 
import pandas as pd 
from tqdm import trange 
import sys
import pickle

from distinctipy import distinctipy

seq_len = 200

device = "cuda:1"

row_length = 9

torch.cuda.set_device(device)

# %%
""" 定义设备和相关参数 """
device = torch.device("cuda" if torch.cuda.is_available() else "cpu" )  
# %%
DATA_DIR =  "dataset"
def inv_trans(x):
    return x

# %%
""" 定义evaluate_agent 函数 """
def evaluate_agent(agent,test_states,test_bm_preds,test_y):
    with torch.no_grad():
        weights = agent.select_action(test_states)

    act_counter = Counter(weights.argmax(1))
    act_sorted = sorted([(k,v) for k,v in act_counter.items()])
    weighted_y = weights * test_bm_preds 
    weighted_y = weighted_y.sum(1)

    mae_loss = mean_absolute_error(inv_trans(test_y),inv_trans(weighted_y))
    mape_loss = mean_absolute_percentage_error(inv_trans(test_y),inv_trans(weighted_y))
    return mae_loss,mape_loss,act_sorted,weighted_y,weights


# %%
""" 定DDPGAgent类 """
class DDPGAgent():
    def __init__(self,use_td,states,obs_dim,act_dim,hidden_dim=256,lr=3e-6,gamma=0.99,tau=0.005):
        self.actor = Actor(obs_dim,act_dim,hidden_dim).to(device)
        self.target_actor = Actor(obs_dim,act_dim,hidden_dim).to(device)
        self.actor_optimizer = torch.optim.Adam(self.actor.parameters(),lr=lr)

        self.critic = Critic(obs_dim,act_dim,hidden_dim).to(device)
        self.target_critic = Critic(obs_dim,act_dim,hidden_dim).to(device)
        self.critic_optimizer = torch.optim.Adam(self.critic.parameters(),lr=lr)
        
        self.states = states

        self.gamma = gamma
        self.tau = tau
        self.use_td = use_td

        for param,target_param in zip(self.actor.parameters(),self.target_actor.parameters()):
            target_param.data.copy_(param.data)

        for param,target_param in zip(self.critic.parameters(),self.target_critic.parameters()):
            target_param.data.copy_(param.data)


    def select_action(self,obs):
        with torch.no_grad():
            action = self.actor(obs).cpu().numpy()
        return softmax(action,axis=1)

    def update(self,
               sampled_obs_idxes,
               sampled_actions,
               sampled_rewards,
               sampled_weights=None):

        batch_obs = self.states[sampled_obs_idxes]

        with torch.no_grad():
            if self.use_td:
                batch_next_obs = self.states[sampled_obs_idxes + 1]
                target_q = self.target_critic(batch_next_obs,F.softmax(self.target_actor(batch_next_obs),dim=1))
                target_q = sampled_rewards + self.gamma * target_q
            else:
                target_q = sampled_rewards
        
        current_q = self.critic(batch_obs,sampled_actions)
        # critic loss
        # F.mse_loss( self.critic(batch_obs,sampled_actions) ,target_q)
        if sampled_weights is None:
            q_loss = F.mse_loss(current_q,target_q)
        else:
            # weighted mse loss
            q_loss = (sampled_weights * (current_q - target_q) ** 2).sum() / sampled_weights.sum()

        self.critic_optimizer.zero_grad()
        q_loss.backward()
        self.critic_optimizer.step()
        #print(f"current_q_2 =  {current_q}")

        # actor loss 
        # F.softmax(self.actor(batch_obs),dim=1).max(dim=1)
        if sampled_weights is None:
            actor_loss = -self.critic(batch_obs,F.softmax(self.actor(batch_obs),dim=1)).mean()
        else:
            actor_loss = -self.critic(batch_obs,F.softmax(self.actor(batch_obs),dim=1))
            actor_loss = (sampled_weights * actor_loss).sum() / sampled_weights.sum()
        self.actor_optimizer.zero_grad()
        
        actor_loss.backward()
        self.actor_optimizer.step()
        
        if self.use_td:
            for param,target_param in zip(self.critic.parameters(),self.target_critic.parameters()):
                target_param.data.copy_(self.tau * param.data + (1- self.tau) * target_param.data)
        for param,target_param in zip(self.actor.parameters(),self.target_actor.parameters()):
            target_param.data.copy_(self.tau * param.data + (1- self.tau) * target_param.data)
        
        return {
            "q_loss":q_loss.item(),
            "pi_loss":actor_loss.item(),
            "current_q":current_q.mean().item(),
            "target_q":target_q.mean().item()
            }

# %%
""" 定义Env环境类 """
class Env():
    def __init__(self,train_error,train_y,train_preds):
        self.error = train_error
        self.bm_preds = train_preds
        self.y = train_y

    def reward_func(self,idx,action):
        if isinstance(action,int):
            tmp = np.zeros(self.bm_preds.shape[1])
            tmp[action] = 1.0
            action = tmp

        weighted_y = np.multiply(action,self.bm_preds[idx])

        weighted_y = np.array([weighted_y.sum(axis=0)])

        new_mape = mean_absolute_percentage_error(inv_trans(self.y[idx]),inv_trans(weighted_y))
        new_mae = mean_absolute_error(inv_trans(self.y[idx]),inv_trans(weighted_y)) 


        new_error = np.array([*self.error[idx],new_mape])
        
        rank = np.where(np.argsort(new_error) == len(new_error) -1 )[0][0]
        return rank,new_mape,new_mae

# %%
""" 定义ReplayBuffer类别 """
class ReplayBuffer:
    def __init__(self,action_dim,max_size=int(1e5)):

        self.max_size = max_size
        self.ptr = 0
        self.size = 0

        self.states = np.zeros((max_size,1),dtype=np.int32)
        self.actions = np.zeros((max_size,action_dim),dtype=np.float32)
        self.rewards = np.zeros((max_size,1),dtype=np.float32)

    def add(self,state,action,reward):
        self.states[self.ptr] = state
        self.actions[self.ptr] = action
        self.rewards[self.ptr] = reward
        self.ptr = (self.ptr + 1) % self.max_size
        self.size = min(self.size + 1 ,self.max_size)

    def sample(self,batch_size=256):
        ind = np.random.randint(self.size,size=batch_size)
        states = self.states[ind].squeeze()
        actions = torch.FloatTensor(self.actions[ind]).to(device)
        rewards = torch.FloatTensor(self.rewards[ind]).to(device)
        return (states,actions,rewards.squeeze())

# %%
""" 定义get_state_weight 函数 """
def get_state_weight(train_error,act_dim):
    L = len(train_error)
    best_model = train_error.argmin(1)
    best_model_counter = Counter(best_model)
    model_weight = {  k:v/L for k,v in best_model_counter.items()}

    for i in range(act_dim):
        if i not in model_weight.keys():
            model_weight[i] = 1/L
    return model_weight

# %%
""" 定义sparse_explore 函数 """
def sparse_explore(obs,act_dim):
    N = len(obs)
    x = np.zeros((N,act_dim))
    randn_idx =  np.random.randint(0,act_dim,size=(N,))
    x[np.arange(N),randn_idx] = 1

    delta = np.random.uniform(0.02,0.1,size=(N,1))
    x[np.arange(N),randn_idx] -= delta.squeeze()

    noise = np.abs(np.random.randn(N,act_dim))
    noise[np.arange(N),randn_idx] = 0
    noise /= noise.sum(1,keepdims=True)
    noise = delta * noise
    sparse_action = x + noise
    return sparse_action

# %%
""" 定义pretrain_actor 函数 """
def pretrain_actor(obs_dim,act_dim,hidden_dim,states,train_error,cls_weights,valid_states,valid_error):
    best_train_model = torch.LongTensor(train_error.argmin(1)).to(device)
    best_valid_model = torch.LongTensor(valid_error.argmin(1)).to(device)

    actor = Actor(obs_dim,act_dim,hidden_dim).to(device)
    best_actor = Actor(obs_dim,act_dim,hidden_dim).to(device)

    cls_weights = torch.FloatTensor( [ 1/cls_weights[w] for w in range(act_dim)]).to(device)

    L = len(states)
    batch_size = 1
    batch_num = int(np.ceil(L / batch_size))
    optimizer = torch.optim.Adam(actor.parameters(),lr=3e-4)
    loss_fn = nn.CrossEntropyLoss() 
    best_acc = 0
    patience = 0
    max_patience = 80

    #for epoch in trange(200,desc="[Pretain]"):
    for epoch in range(2000):

        epoch_loss = []
        shuffle_idx = np.random.permutation(np.arange(L))
        for i in range(batch_num):
            batch_idx = shuffle_idx[i * batch_size:(i+1)* batch_size]
            optimizer.zero_grad()
            batch_out = actor(states[batch_idx])
            loss = loss_fn(batch_out,best_train_model[batch_idx])
            loss.backward()

            optimizer.step()
            epoch_loss.append(loss.item())
        with torch.no_grad():
            pred = actor(valid_states)
            pred_idx = pred.argmax(1)
            acc = (pred_idx ==  best_valid_model).sum() / len(pred)

        print(f"# epoch {epoch + 1} : loss = {np.average(epoch_loss):.5f}  \t acc = {acc:.3f} patience = {patience}")

        if acc > best_acc:
            best_acc = acc
            patience = 0

            for param,target_param in zip(actor.parameters(),best_actor.parameters()):
                target_param.data.copy_(param.data)
        else:
            patience += 1
        if patience == max_patience:
            break
    with torch.no_grad():
        pred = best_actor(valid_states)
        pred_idx = pred.argmax(1)
        acc = (pred_idx ==  best_valid_model).sum() / len(pred)

    print(f"valid acc for pretrained actor:{acc:.3f}")
    return best_actor

# %%
""" 随机从所有电池中进行采样 ,强化学习从训练电池中选择 """

""" 
import re 
pattern = "^[0-9].*csv$"
test_cols = []

for file in os.listdir("./basemodels/model_all/"):
    results = re.findall(pattern,file)
    if len(results) != 0 :
        test_cols.append(file[:-4])
       
# 
        
# temp_cols = list(pd.read_csv("../dataset/processed_data/shu_old_fade_rate.csv",index_col=0).columns)

temp_cols = list(pd.read_csv("../dataset/processed_data/nature_processed_dup.csv",index_col=0).columns)
train_cols = [ x for x in temp_cols if x not in test_cols]

"""



import pickle 



# %%
# 划分
import random 




"""  
test_cols_length = len(test_cols)
train_rate = 0.2

train_weight_length  = int(train_rate * test_cols_length)
train_weight_cols = random.sample(test_cols,train_weight_length)

test_weight_cols = [ x for x in test_cols if x not in train_weight_cols]

"""

# %%
import re 

train_battery_num = 0
for file in os.listdir("./basemodels/remove_model"):
    train_battery_num += 1

color_lst = distinctipy.get_colors(train_battery_num)


pattern = "^[0-9].*xlsx$"
temp_list = []

for file in os.listdir("./basemodels/model_all/"):
    results = re.findall(pattern,file)
    if len(results) != 0 :
        temp_list.append(file[:-5])


train_cols = temp_list[:5]

pattern = "^[0-9].*csv$"
temp_list = []

for file in os.listdir("./basemodels/model_all/"):
    results = re.findall(pattern,file)
    if len(results) != 0 :
        temp_list.append(file[:-4])

test_cols = temp_list

test_weight_cols = test_cols
train_weight_cols = train_cols

print(f"train_cols: {train_cols}")
print(f"test_cols: {test_cols}")

# %%
""" 定义split_x_y函数 """
def split_x_y(series,context_length=120,pred_length=24):
    series_length = len(series)

    x = []
    y = []

    for i in range(series_length - context_length - pred_length):
        x.append(series[i:i+context_length])
        y.append(series[i+context_length:i+context_length + pred_length])
        
    return np.array(x),np.array(y)

# %%
global_train_weights = train_weight_cols
global_test_weights = test_weight_cols
global_rm_is_accuracy = True
global_rm_is_complex = True



# %%
''' train_rlmc的训练 '''
def train_rlmc(cur_idx):

    
    
    # 将用于rlmc训练的电池改为固定的电池
    train_weight_cols = global_train_weights
    test_weight_cols = global_test_weights

    # 每次取不同的train_weight
    train_weight_cols = train_weight_cols[cur_idx:cur_idx + 1] 
    
    print(f"train_weight_cols = {train_weight_cols}")

    #arg1 = int(sys.argv[1])
    arg1 = 1
    dataset_path = f"./dataset_{arg1}"
    result_path = f"./results/data_{arg1}"

    if not os.path.exists(dataset_path):
        os.makedirs(dataset_path)

    if not os.path.exists(result_path):
        os.makedirs(result_path)
        
        
    
    
    # 构造训练数据

    for idx,col in enumerate(train_weight_cols):
        # 读取真实值
        col_real =  pd.read_excel(f"./basemodels/model_all/{col}.xlsx")["ture"][seq_len:]
        #col_real =  pd.read_csv("../dataset/processed_data/shu_old_fade_rate.csv")["10A_(2)"].dropna()[:-49][seq_len:]
        
        
        # 划分x和y
        context_length = 10
        predict_length = 1
        col_x,col_y = split_x_y(col_real,context_length,predict_length)

        col_x = np.expand_dims(col_x,2)

        if idx == 0:
            all_col_x = col_x
            all_col_y = col_y
        else:
            all_col_x = np.concatenate((all_col_x,col_x),axis=0)
            all_col_y = np.concatenate((all_col_y,col_y),axis=0)
        
        # preds 
        preds_dict = {}
        for i in range(train_battery_num ):
            preds_dict[f"remove_{i}"] = pd.read_excel(f"./basemodels/remove_model/model_remove{i}/{col}.xlsx")["predict"][seq_len:][context_length+predict_length:]

        preds_df = pd.DataFrame.from_dict(preds_dict)
        model_all_real_df = pd.read_excel(f"./basemodels/model_all/{col}.xlsx")["ture"][seq_len:][context_length+predict_length:]
        
        if idx == 0:
            all_preds_df = preds_df
            total_model_all_real_df = model_all_real_df
        else:
            all_preds_df = pd.concat([all_preds_df,preds_df])
            total_model_all_real_df = pd.concat([total_model_all_real_df,model_all_real_df])

        # errors
        col_real = col_y
        errors_dict = {}
        for i in range(train_battery_num ):
            temp_predict =  pd.read_excel(f"./basemodels/remove_model/model_remove{i}/{col}.xlsx")["predict"][seq_len:][context_length+predict_length:].to_numpy().reshape(-1,1)
            errors = []

            for j in range(len(col_real)):
                errors.append(  mean_absolute_error(col_real[j],temp_predict[j])) 
            errors = np.array(errors)

            errors_dict[f"remove_{i}"] = errors

        errors_df = pd.DataFrame.from_dict(errors_dict)
        if idx == 0:
            all_errors_df = errors_df
        else:
            all_errors_df = pd.concat([all_errors_df,errors_df])

    all_preds_df.to_csv(f"{dataset_path}/train_preds.csv")
    total_model_all_real_df.to_csv(f"{dataset_path}/train_model_all_real.csv")
    all_errors_df.to_csv(f"{dataset_path}/train_errors.csv")

    with open(f"{dataset_path}/train_x.npy","wb") as f:
        np.save(f,all_col_x)
    with open(f"{dataset_path}/train_y.npy","wb") as f:
        np.save(f,all_col_y)


    """ RLMC的参数"""
    base_path = f"{result_path}"

    use_weight = False
    use_td = True
    use_extra = True
    use_pretrain = True
    epsilon = 0.5


    """读取train_x，train_y等结果"""

    train_x = np.load(f"{dataset_path}/train_x.npy")
    valid_x = train_x    

    train_y = np.load(f"{dataset_path}/train_y.npy")
    valid_y = train_y

    train_error = pd.read_csv(f"{dataset_path}/train_errors.csv",index_col=0).to_numpy()
    valid_error = train_error

    train_preds = pd.read_csv(f"{dataset_path}/train_preds.csv",index_col=0).to_numpy()
    valid_preds = train_preds

    train_x = np.swapaxes(train_x,2,1)
    valid_x = np.swapaxes(valid_x,2,1)

    L = len(train_x) - 1 if use_td else len(train_x)

    FEAT_LEN = 20   

    train_x = train_x[:,:,-FEAT_LEN:]
    valid_x = valid_x[:,:,-FEAT_LEN:]

    states = torch.FloatTensor(train_x).to(device)
    valid_states = torch.FloatTensor(valid_x).to(device)

    cur_agent_states_max = states[:,0,0].detach().cpu().numpy().max()
    cur_agent_states_min = states[:,0,0].detach().cpu().numpy().min()

    """ 构建obs_dim 和 act_dim 参数 构建env参数"""
    obs_dim = train_x.shape[1]
    act_dim = train_error.shape[-1]
    env = Env(train_error,train_y,train_preds)


    """" 构建batch_buffer,q_mae和get_batch_*_reward函数等等 """

    best_model_weight = get_state_weight(train_error,act_dim=act_dim)

    if not os.path.exists(f"{dataset_path}/batch_buffer.csv"):
        batch_buffer = []
        for state_idx in trange(L,desc="[Create buffer]"):
            best_model_idx = train_error[state_idx].argmin()
            for action_idx in range(act_dim):
                rank,mape,mae = env.reward_func(state_idx,action_idx)
                batch_buffer.append((state_idx,action_idx,rank,mape,mae,best_model_weight[best_model_idx]))
        batch_buffer_df = pd.DataFrame(batch_buffer,columns=["state_idx","action_idx","rank","mape","mae","weight"])
        batch_buffer_df.to_csv(f"{dataset_path}/batch_buffer.csv")
    else:
        batch_buffer_df = pd.read_csv(f"{dataset_path}/batch_buffer.csv",index_col=0)
        
    reward_dim = train_battery_num  
    quantile_rate = 1.0/(reward_dim+1)

    q_mae = [batch_buffer_df['mae'].quantile(quantile_rate * i ) for i in range(1,reward_dim+1) ]

    if use_td:
        batch_buffer_df = batch_buffer_df.query(f"state_idx < {L}")

    def get_mae_reward(q_mae,mae,R=1):
        q = 0
        while (q < reward_dim) and (mae > q_mae[q]):
            q += 1 
        reward = -R + 2 * R * (reward_dim-q)/reward_dim
        return reward

    def get_rank_reward(rank,R=1):
        reward = -R + 2 * R * (reward_dim-rank)/reward_dim
        return reward

    def get_batch_rewards(env,idxes,actions):
        rewards = []
        mae_list = []
        for i in range(len(idxes)):
            rank,new_mape,new_mae = env.reward_func(idxes[i],actions[i])

            rank_reward = get_rank_reward(rank,1)
            mae_reward = get_mae_reward(q_mae,new_mae,1)

            combined_reward = mae_reward + rank_reward
            mae_list.append(new_mae)
            rewards.append(rank_reward)
        return rewards,mae_list
    
    
    
    
    def get_batch_rewards_rm(env,idxes,actions):
        global global_rm_is_accuracy 
        global global_rm_is_complex
        # 增加reward_machine之后的奖励函数
        rewards = []
        mae_list = []
        for i in range(len(idxes)):
            
            # 获取rank值 和mae 
            rank,new_mape,new_mae = env.reward_func(idxes[i],actions[i])

            # 计算RM的准确度奖励值
            # 好坏值的计算方式对如果结果排名前50%则，结果视为好的，否则视为坏的
            accuracy_rate = rank * 1.0 / reward_dim
            accuracy_threshold = 0.5 
            
            accuracy_reward = 0
            
            # 准确度和复杂性比例
            rate_for_accuracy_and_complexity = 0.8
            
            if accuracy_rate > accuracy_threshold:
                # 准确
                if global_rm_is_accuracy : 
                    # 之前是准确的
                    accuracy_reward += 0
                else:
                    # 之前是不准确的
                    accuracy_reward += rate_for_accuracy_and_complexity 
                    global_rm_is_accuracy = True  
            else: 
                # 不准确
                if global_rm_is_accuracy : 
                    # 之前是准确的
                    accuracy_reward += -rate_for_accuracy_and_complexity 
                    global_rm_is_accuracy = False
                else:
                    # 之前是不准确的
                    accuracy_reward += -rate_for_accuracy_and_complexity * 0.5

            
            # 复杂性奖励
            complexity_reward = 0
            
            actions_after_sorted = sorted(actions[i])
            first_action,second_action = actions_after_sorted[0],actions_after_sorted[1]
            
            dist_action = first_action - second_action 
            complexity_threshold = 0.75
            
            if dist_action < complexity_threshold:
                # 复杂
                if global_rm_is_complex : 
                    # 以前是复杂
                    complexity_reward += 0
                else: 
                    # 以前是单一
                    complexity_reward += (1 - rate_for_accuracy_and_complexity)
                    global_rm_is_complex = True
            else: 
                # 单一
                if global_rm_is_complex:
                    # 以前是复杂
                    complexity_reward += - (1 - rate_for_accuracy_and_complexity)
                    global_rm_is_complex = False
                else: 
                    # 以前是单一
                    complexity_reward += - (1 - rate_for_accuracy_and_complexity) * 0.5

            mae_list.append(new_mae)            
            rewards.append(accuracy_reward + complexity_reward)
        return rewards,mae_list

    def get_batch_mae_rewards(env,idxes,actions):
        rewards = []
        for i in range(len(idxes)):
            rank,new_mape,new_mae = env.reward_func(idxes[i],actions[i])
            mae_reward = get_mae_reward(q_mae,new_mae,1)
            rewards.append(mae_reward)
        return rewards

    def get_batch_rank_rewards(env,idxes,actions):
        rewards = []
        for i in range(len(idxes)):
            rank,new_mape,new_mae = env.reward_func(idxes[i],actions[i])
            rank_reward = get_rank_reward(rank,1)
            rewards.append(rank_reward)
        return rewards



    """构建agent和replay_buffer"""
    state_weights = [1/best_model_weight[i]  for i in train_error.argmin(1)]
    if use_weight:
        state_weights = torch.FloatTensor(state_weights).to(device)
    else:
        state_weights = None
    
    

    agent = DDPGAgent(use_td,states,obs_dim,act_dim,hidden_dim=100)
    replay_buffer = ReplayBuffer(act_dim,max_size = int(1e5))
    extra_buffer = ReplayBuffer(act_dim,max_size=int(1e5))

    threshold_for_buffer = 0.8   
    large_buffer = ReplayBuffer(act_dim,max_size = int(1e5))
    small_buffer = ReplayBuffer(act_dim,max_size = int(1e5))

    train_actions = agent.select_action(states)
    train_idxes = np.arange(L)

    print("=========initialed=========")

    """ 进行预训练 """ 
    if use_pretrain:
        pretrained_actor = pretrain_actor(
            obs_dim,
            act_dim,
            hidden_dim=100,
            states = states,
            train_error = train_error,
            cls_weights = best_model_weight,
            valid_states = valid_states,
            valid_error = valid_error
        )
        
        for param,target_param in zip(pretrained_actor.parameters(),agent.actor.parameters()):
            target_param.data.copy_(param.data)

        for param,target_param in zip(pretrained_actor.parameters(),agent.target_actor.parameters()):
            target_param.data.copy_(param.data)
            

    
    train_idxes = np.arange(L)
    train_actions = agent.select_action(states)


    # 打印初始化agent之后的reward值
    plt.figure()
    plt.title("after pretrain all_reward")
    init_rewards,_ = get_batch_rewards_rm(env,train_idxes,train_actions)
    plt.plot(init_rewards,label="all")
    plt.show()
    plt.savefig("./temp/after_pretrained_all_reward.png") 

    # 打印初始化之后的agent扽weights值
    test_mae_loss,test_mape_loss,count_list,weighted_y,weights = evaluate_agent(agent,states,train_preds,train_y)   
    plt.figure()
    plt.title(f"after pretrained  weights plot")
    for i in range(train_battery_num ):
        plt.plot(weights[:,i],label=f"{i}",color=color_lst[i])
    plt.legend()
    plt.savefig("./temp/after_pretrained_weights_plot.png") 

    print(f"test_mae_loss:{test_mae_loss}")  

    """对于replay_buffer进行warm up"""
    for _ in trange(1,desc="[Warm up]"):
        shuffle_idxes = np.random.randint(0,L,30)
        sampled_states = states[shuffle_idxes]
        sampled_actions = agent.select_action(sampled_states)
        sampled_rewards,_ = get_batch_rewards_rm(env,shuffle_idxes,sampled_actions)
        for i in range(len(sampled_states)):
            replay_buffer.add(shuffle_idxes[i],sampled_actions[i],sampled_rewards[i])
            current_states = states[shuffle_idxes[i]]
            if current_states[0,0].item() > threshold_for_buffer:
                large_buffer.add(shuffle_idxes[i],sampled_actions[i],sampled_rewards[i])
            else:
                small_buffer.add(shuffle_idxes[i],sampled_actions[i],sampled_rewards[i])
            if use_extra and sampled_rewards[i] <= -1:
                extra_buffer.add(shuffle_idxes[i],sampled_actions[i],sampled_rewards[i])

    """ 进行RLMC算法的训练 """
    # 记录rlmc训练过程中的最佳的actor表现
    img_path = "./results/temp"
    best_actor = Actor(obs_dim,act_dim,hidden_dim=100).to(device)
    for param,target_param in zip(agent.actor.parameters(),best_actor.parameters()):
        target_param.data.copy_(param.data)

    # rlmc训练过程的参数
    step_size = 400
    step_num = int(np.ceil(L / step_size))  
    best_mape_loss = np.inf
    patience,max_patience= 0,5

    # 记录训练过程中每个epoch 的agent方便后续测试
    epoch_agent_list = []
    epoch_idx_list = [ x for x in range(200)]
    
    # 通过多个epoch的循环进行训练
    for epoch in trange(12,desc="epoch:"):
        
        t1 = time.time()
        
        valid_mae_loss,valid_mape_loss,count_list,weighted_y,weights = evaluate_agent(agent,valid_states,valid_preds,valid_y)

        print(f"\n # Epoch {epoch + 1} ({( time.time()-t1)/60 :.2f} min):" 
                f"valid_mae_loss : {valid_mae_loss:.3f}\t" 
                f"valid_mape_loss : {valid_mape_loss * 100 :.3f} \t"

                )

         
        if valid_mape_loss < best_mape_loss:
            best_mape_loss = valid_mape_loss
            patience = 0

            for param,target_param in zip(agent.actor.parameters(),best_actor.parameters()):
                target_param.data.copy_(param.data)
        else:
            patience += 1
            
        if patience == max_patience:
            break
        
        
        q_loss_lst,pi_loss_lst,q_lst,target_q_lst = [],[],[],[]
        shuffle_idx = np.random.permutation(np.arange(L))
        
        for i in trange(step_num):
            print(f"step {i}/{step_num}")
        
            batch_idx = shuffle_idx[i * step_size: (i+1)* step_size]
            batch_states = states[batch_idx]

            if np.random.random() < epsilon:
                batch_actions = sparse_explore(batch_states,act_dim)
            else:
                batch_actions = agent.select_action(batch_states)
            batch_rewards,batch_mae = get_batch_rewards_rm(env,batch_idx,batch_actions)
            for j in range(len(batch_idx)):
                replay_buffer.add(batch_idx[j],batch_actions[j],batch_rewards[j])
                if use_extra and batch_rewards[j] <= -1:
                    extra_buffer.add(batch_idx[j],batch_actions[j],batch_rewards[j])        

            sample_size = 320
            sample_rate = 0.1
            large_size = int(sample_size * sample_rate)
            small_size = sample_size - large_size
            
            for step_idx in range(3):
                
                """
                sampled_obs_idxes,sampled_actions,sampled_rewards = large_buffer.sample(large_size)
                small_sampled_obs_idxes,small_sampled_actions,small_sampled_rewards = small_buffer.sample(small_size)
                sampled_obs_idxes = np.concatenate((sampled_obs_idxes,small_sampled_obs_idxes),axis=0)
                sampled_actions = torch.cat((sampled_actions,small_sampled_actions),0)
                sampled_rewards = torch.cat((sampled_rewards,small_sampled_rewards),0)
                """
                
                sampled_obs_idxes,sampled_actions,sampled_rewards = replay_buffer.sample(sample_size)
                if use_weight:
                    sampled_weights = state_weights[sampled_obs_idxes]
                else:
                    sampled_weights = None
                
                train_idxes = np.arange(L)
                train_actions = agent.select_action(states)

                """ 
                plt.figure()
                plt.title(f" before reward")
                all_rewards,_ = get_batch_rewards(env,train_idxes,train_actions)
                plt.plot(all_rewards,label="all")
                plt.savefig(f"{img_path}/{epoch}_{i}_{step_idx}_before_reward.png")
                plt.show()
                
                test_mae_loss,test_mape_loss,count_list,weighted_y,weights = evaluate_agent(agent,states,train_preds,train_y)  
                plt.figure()
                plt.title(f"before weights plot")
                for j in range(train_battery_num ):
                    plt.plot(train_actions[:,j],label=f"{j}")
                plt.legend()
                plt.savefig(f"{img_path}/{epoch}_{i}_{step_idx}_before_weights.png")
                plt.show()

                q_list = agent.critic(states,F.softmax(agent.actor(states),dim=1) )


                plt.figure()
                plt.title("q_list")
                plt.plot(q_list.detach().cpu().numpy(),label=f"actions")
                for j in range(train_battery_num ):
                    temp = np.zeros(train_battery_num )
                    temp[0] = j
                    q_1_action = [ temp  for _ in range(len(states))    ]


                    q_1_action = torch.Tensor(q_1_action).to(device)
                    q_1 = agent.critic(states,q_1_action).detach().cpu().numpy()

                    plt.plot(q_1,label=f"{j}")
                
                
                plt.legend()
                plt.show()
                
                """
                info = agent.update(sampled_obs_idxes,sampled_actions,sampled_rewards,sampled_weights)    
                    
                """
                train_idxes = np.arange(L)
                train_actions = agent.select_action(states)
                
                plt.figure()
                plt.title(f" after  all_reward")
                all_rewards,_ = get_batch_rewards(env,train_idxes,train_actions)
                plt.plot(all_rewards,label="all")
                plt.savefig(f"{img_path}/{epoch}_{i}_{step_idx}_after_reward.png")
                
                plt.figure()
                plt.title(f"after  weights plot")
                for j in range(train_battery_num ):
                    plt.plot(train_actions[:,j],label=f"{j}")
                plt.legend()
                plt.savefig(f"{img_path}/{epoch}_{i}_{step_idx}_after_weights.png")
                """
                
                pi_loss_lst.append(info["pi_loss"])
                q_loss_lst.append(info["q_loss"])
                q_lst.append(info['current_q'])
                target_q_lst.append(info['target_q'])

                #print(f"i={i} step_idx = {step_idx}  = {info['q_loss']}")
        
                """  
                print(f"\n # step {step_idx} " 
                f"pi_loss : {info['pi_loss']:.3f}\t" 
                f"q_loss : {info['q_loss']:.5f} \t"
                f"current_q : {info['current_q']:.5f} \t"
                f"target_q:{info['target_q']:.5f} \n"
                )    
                """
                use_extra = False
            
                if use_extra and extra_buffer.ptr > sample_size:
                    sampled_obs_idxes,sampled_actions,sampled_rewards = extra_buffer.sample(sample_size)
                    if use_weight:
                        sampled_weights = state_weights[sampled_obs_idxes]
                    else:
                        sampled_weights = None

                    info = agent.update(sampled_obs_idxes,sampled_actions,sampled_rewards,sampled_weights)
                    pi_loss_lst.append(info['pi_loss'])
                    q_loss_lst.append(info['q_loss'])
                    q_lst.append(info['current_q'])
                    target_q_lst.append(info['target_q'])

            test_mae_loss,test_mape_loss,count_list,weighted_y,weights = evaluate_agent(agent,states,train_preds,train_y)   
            train_idxes = np.arange(L)
            train_actions = agent.select_action(states)
            
            """  
            plt.figure()
            plt.title(f" after trained {i} all_reward")
            all_rewards,_ = get_batch_rewards(env,train_idxes,train_actions)
            plt.plot(all_rewards,label="all")
            plt.close()
            
            plt.figure()
            plt.title(f"after trained {i}  weights plot")
            for j in range(train_battery_num ):
                plt.plot(weights[:,j],label=f"{j}")
            plt.legend()
            plt.close()    
            """


        
            
        if epoch  in epoch_idx_list :
            temp_agent = DDPGAgent(use_td,states,obs_dim,act_dim,hidden_dim=100)
        
            for param,target_param in zip(agent.actor.parameters(),temp_agent.actor.parameters()):
                target_param.data.copy_(param.data)

            for param,target_param in zip(agent.target_actor.parameters(),temp_agent.target_actor.parameters()):
                target_param.data.copy_(param.data)
                
            epoch_agent_list.append(temp_agent)
        
        epsilon = max(epsilon-0.2,0.1)

        train_idxes = np.arange(L)
        train_actions = agent.select_action(states)
        
        plt.figure()
        plt.title(f" after  all_reward")
        all_rewards,_ = get_batch_rewards_rm(env,train_idxes,train_actions)
        plt.plot(all_rewards,label="all")
        plt.show()
        plt.savefig(f"./temp/{epoch}_after_all_reward.png") 

        plt.figure()
        plt.title(f"after  weights plot")
        for j in range(train_battery_num ):
            plt.plot(train_actions[:,j],label=f"{j}",color=color_lst[j])
        plt.legend()
        plt.show()
        plt.savefig(f"./temp/{epoch}_after_all_weights.png") 
    
    for param,target_param in zip(agent.actor.parameters(),best_actor.parameters()):
        param.data.copy_(target_param.data)
    
    
    return agent,train_actions,(cur_agent_states_max,cur_agent_states_min)



# %%

agent_list = []

states_max_list = []
states_min_list = []

for i in range(len(train_weight_cols)):
    isLocalOptim = True 
    max_local_iter = 5

    while isLocalOptim:
        temp_agent,weights,(cur_max,cur_min) = train_rlmc(i)
        
        temp_weights = weights.argmax(axis=1)
        isLocalOptim = True
        for idx in range(1,len(temp_weights)):
            if temp_weights[idx] != temp_weights[0]:
                isLocalOptim = False
        max_local_iter -= 1 
        
        if max_local_iter <= 0 :
            break
    agent_list.append(temp_agent)
    states_max_list.append(cur_max)
    states_min_list.append(cur_min)


# %%
arg1 = 1
dataset_path = f"./dataset_{arg1}"
result_path = f"./results/data_{arg1}"
base_path = f"{result_path}"

# %%
with open(f"{result_path}/agent_list.pkl","wb") as handle:
    pickle.dump(agent_list,handle)
    
with open(f"{result_path}/states_max_list.pkl","wb") as handle:
    pickle.dump(states_max_list,handle)

with open(f"{result_path}/states_min_list.pkl","wb") as handle:
    pickle.dump(states_min_list,handle)

# %%
with open(f"{result_path}/agent_list.pkl","rb") as handle:
    agent_list = pickle.load(handle)
    
    
with open(f"{result_path}/states_max_list.pkl","rb") as handle:
    states_max_list = pickle.load(handle)
    
    
with open(f"{result_path}/states_min_list.pkl","rb") as handle:
    states_min_list = pickle.load(handle)

# %%
""" 在rlmc训练完成之后对于结果进行测试 """

test_states_list = []
test_preds_list = []
test_y_real_list = []
model_all_pred_list = []
remove_list = []
epoch_mae_list = [] 

mae_dict = {}
mae_weighted_list = []
mae_all_list = []
best_basemodel_list = []

arg1 = 1
dataset_path = f"./dataset_{arg1}"
result_path = f"./results/data_{arg1}"
base_path = f"{result_path}"

FEAT_LEN = 20 

weighted_y_dict = {}
test_y_real_dict = {}
model_all_pred_dict = {}

for i in range(train_battery_num ):
    remove_list.append([])

for col in test_weight_cols:
    print(col)
#for col in train_weight_cols:
    
    col_real =  pd.read_csv(f"./basemodels/model_all/{col}.csv")["ture"][seq_len:]
    col_model_all_predict = pd.read_csv(f"./basemodels/model_all/{col}.csv")["predict"][seq_len:]

    # 划分x和y
    context_length = 10
    predict_length = 1
    
    _,test_y_real = split_x_y(col_real,context_length,predict_length)

    col_x,col_y = split_x_y(col_model_all_predict,context_length,predict_length)

    # save x y 
    col_x = np.expand_dims(col_x,2)
    
    with open(f"{dataset_path}/test_x.npy","wb") as f:
        np.save(f,col_x)

    with open(f"{dataset_path}/test_y.npy","wb") as f:
        np.save(f,col_y)

    preds_dict = {}

    for i in range(train_battery_num ):
        preds_dict[f"remove_{i}"] = pd.read_csv(f"./basemodels/remove_model/model_remove{i}/{col}.csv")["predict"][seq_len:][context_length+predict_length:]

    preds_df = pd.DataFrame.from_dict(preds_dict) 

    preds_df.to_csv(f"{dataset_path}/test_preds.csv")

    col_real = col_y 

    pd.read_csv(f"./basemodels/model_all/{col}.csv")["predict"][seq_len:][context_length+predict_length:].to_csv(f"{dataset_path}/test_model_all_pred.csv")

    errors_dict = {}

    for  i in range(train_battery_num ):
        temp_predict = pd.read_csv(f"./basemodels/remove_model/model_remove{i}/{col}.csv")["predict"][seq_len:][context_length+predict_length:].to_numpy().reshape(-1,1)

        errors = []

        for j in range(len(col_real)):
            errors.append(mean_squared_error(col_real[j],temp_predict[j]))

        errors = np.array(errors)
        errors_dict[f"remvoe_{i}"] = errors 

    errors_df = pd.DataFrame.from_dict(errors_dict)
    errors_df.to_csv(f"{dataset_path}/test_errors.csv")

    test_x = np.load(f"{dataset_path}/test_x.npy")
    test_y = np.load(f"{dataset_path}/test_y.npy")
    
    test_error = pd.read_csv(f"{dataset_path}/test_errors.csv",index_col=0).to_numpy()
    test_preds = pd.read_csv(f"{dataset_path}/test_preds.csv",index_col=0).to_numpy()
    test_x = np.swapaxes(test_x,2,1)

    def mae_func(y,y_):
        results = []
        for i in range(len(y)):
            results.append(abs(y[i] - y_[i]))
        return results

    def plot_mae(start,end):

        plt.figure()
        plt.title(f"{col}_{start} - {end} mae  plots")

        for i in range(start,end):
            results = mae_func(test_y,test_preds[:,i])
            plt.plot(results,label=i)
        plt.legend()           
        plt.show()

    
    test_x = test_x[:,:,-FEAT_LEN:]

    test_states = torch.FloatTensor(test_x).to(device)

    def plot_weights(msg):
        plt.figure()
        plt.title(f"{msg}")
        for i in range(train_battery_num ):
            plt.plot(weights[:,i],label=f"{i}",color=color_lst[i])
        plt.legend()
        plt.savefig(f"{base_path}/weights/{msg}.png")
        plt.show()
    
    cur_states_max = test_states[:,0,0].detach().cpu().numpy().max()
    cur_states_min = test_states[:,0,0].detach().cpu().numpy().min()
    
    # 决定使用哪个agent
    min_test_mae_loss = float('inf')
    
    agent = agent_list[0]
    
    for temp_idx,temp_agent in enumerate(agent_list):
        
        test_mae_loss,_,_,_,weights = evaluate_agent(temp_agent,test_states,test_preds,col_y)
        
        cur_agent_states_max = states_max_list[temp_idx]
        cur_agent_states_min = states_min_list[temp_idx]
        
        diff_1 = abs(cur_states_min - cur_agent_states_max)
        diff_2 = abs(cur_states_max - cur_agent_states_min)

        diff_threshhold = 0.1
        
        
        # 没有重叠则跳过
        if  cur_states_max < cur_agent_states_min  or cur_states_min > cur_agent_states_max :
            continue
            pass
            
        # 重叠区域过小跳过
        if diff_1 < diff_threshhold or diff_2 < diff_threshhold: 
            continue
            pass
                
        #plot_weights(f"[{temp_idx}] _weights loss = {test_mae_loss}")
        
        if min_test_mae_loss > test_mae_loss:
            
            
            min_test_mae_loss = test_mae_loss
            agent = temp_agent
            
    test_mae_loss,test_mape_loss,count_list,weighted_y,weights = evaluate_agent(agent,test_states,test_preds,test_y_real)   

    if not os.path.exists(f"{base_path}/weights"):
        os.makedirs(f"{base_path}/weights")
        
    plot_weights(f"{col} weights ")

    model_all_pred = pd.read_csv(f"{dataset_path}/test_model_all_pred.csv",index_col=0)
    
    model_all_pred_list.append(model_all_pred)
    all_mae =   mean_absolute_error(test_y_real,model_all_pred)

    mae_weighted_list.append(test_mae_loss)
    
    mae_all_list.append(all_mae)    

    min_mae = float('inf')

    for j in range(train_battery_num ):
        current_mae = mean_absolute_error(test_y_real,test_preds[:,j])          
        if current_mae < min_mae :
            min_mae = current_mae

        remove_list[j].append(current_mae)

    best_basemodel_list.append(min_mae)
    
    test_states_list.append(test_states)
    test_preds_list.append(test_preds)
    test_y_real_list.append(test_y_real)
    
    weighted_y_dict[col] = weighted_y
    test_y_real_dict[col] = test_y_real.squeeze()
    model_all_pred_dict[col] = model_all_pred.to_numpy().squeeze()
    


# xlsx文件每行的最小值标黄
import openpyxl 
from openpyxl.styles import PatternFill 


def color_min(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb["Sheet1"]
    
     
    col_list = ["B","C"]
    
    
    for row in range(2,2 + row_length):
        min_value = float('inf')
        min_index = ""

        for col in col_list:
            current_index = f"{col}{row}"
            current_value =  ws[current_index].value

            if current_value < min_value:
                min_value = current_value
                min_index = current_index

        ws[min_index].fill = PatternFill(patternType='solid',fgColor="00FFFF00")
    wb.save(filename)
    
def color_max(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb["Sheet1"]
    
     
    col_list = ["B","C"]
    
    
    for row in range(2,2 + row_length):
        max_value = float('-inf')
        max_index = ""

        for col in col_list:
            current_index = f"{col}{row}"
            current_value =  ws[current_index].value

            if current_value > max_value:
                max_value = current_value
                max_index = current_index

        ws[max_index].fill = PatternFill(patternType='solid',fgColor="00FFFF00")
    wb.save(filename)



def color_smaller_than_last_column(filename):
    wb = openpyxl.load_workbook(f"{base_path}/mae.xlsx")
    ws = wb["Sheet1"]

     
    last_col = "C"
    col_length = 2

    for row in range(2,2 + row_length):
        min_value = ws[f"{last_col}{row}"].value

        for col in [ chr(x)  for x in range(66 , 66+col_length)]:
            current_index = f"{col}{row}"
            current_value =  ws[current_index].value

            if current_value < min_value:
                ws[current_index].fill = PatternFill(patternType='solid',fgColor="00FFFF00") 

# 保存原始预测结果
pred_value_path = f"{result_path}/pred_value"

if not os.path.exists(pred_value_path):
    os.mkdir(pred_value_path)
   
with open(f"{pred_value_path}/weighted_y_dict.pkl","wb") as handler: 
    pickle.dump(weighted_y_dict,handler)
    
with open(f"{pred_value_path}/test_y_real_dict.pkl","wb") as handler: 
    pickle.dump(test_y_real_dict,handler)
    
with open(f"{pred_value_path}/model_all_pred_dict.pkl","wb") as handler: 
    pickle.dump(model_all_pred_dict,handler)
    
#  计算mae
def metrics_func(metrics_func=mean_absolute_error,name="mae",is_min=True):
    columns = list(weighted_y_dict.keys())

    total_col_mae =  []

    my_pred_list = []
    all_pred_list = []
    real_value_list = []

    for col in columns : 

        # 分电池计算
        col_mae = []
        my_pred = weighted_y_dict[col]
        all_pred = model_all_pred_dict[col]
        real_value = test_y_real_dict[col]
        Informer_mae = metrics_func(real_value,all_pred)
        RLInformer_mae = metrics_func(real_value,my_pred)
        
        col_mae.append(Informer_mae)
        col_mae.append(RLInformer_mae)
        
        total_col_mae.append(col_mae)
        
        # 总体计算
        my_pred_list = my_pred_list + list(weighted_y_dict[col])
        all_pred_list = all_pred_list + list(model_all_pred_dict[col])
        real_value_list = real_value_list + list(test_y_real_dict[col])
        
        
    avg_Informer_mae = metrics_func(real_value_list,all_pred_list)
    avg_RLInformer_mae = metrics_func(real_value_list,my_pred_list)

    total_col_mae.append([avg_Informer_mae,avg_RLInformer_mae])

    total_col_mae_df = pd.DataFrame(total_col_mae)
    total_col_mae_df.columns = ["Informer","RLInformer"]
    total_col_mae_df.index = columns + ["AVG"]

    total_col_mae_df.to_excel(f"{base_path}/{name}.xlsx")
    


# %%
metrics_func(mean_squared_error,"mse")
metrics_func(mean_absolute_error,"mae")
metrics_func(mean_absolute_percentage_error,"mape")
metrics_func(r2_score,"r2",False)


